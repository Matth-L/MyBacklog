"""Imports data from the CSV / XLSX files in the 'My Backlog' format."""
import re
import json
import unicodedata
import difflib
import pandas as pd
from .db import get_conn
from . import sequel_guard
from .dateutils import month_name_to_number


def _as_year(val):
    """Tries to interpret a value as a year (1900-2100). Handles the case
    where pandas converted a whole-number column to float because of a NaN
    elsewhere (e.g. 2024.0 instead of 2024)."""
    if pd.isna(val):
        return None
    try:
        f = float(str(val).strip())
    except (TypeError, ValueError):
        return None
    if f.is_integer() and 1900 <= f <= 2100:
        return int(f)
    return None


def _norm(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9]+", " ", s).strip()
    return s


def _to_bool(v):
    if pd.isna(v):
        return None
    s = str(v).strip().lower()
    if s in ("yes", "true", "oui", "1"):
        return 1
    if s in ("no", "false", "non", "0"):
        return 0
    return None


def _to_float(v):
    try:
        if pd.isna(v):
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _to_int(v):
    f = _to_float(v)
    return int(f) if f is not None else None


def _find_sheet(xl, *candidates):
    """Trouve une feuille par nom approximatif (insensible aux accents/casse)."""
    norm_map = {_norm(name): name for name in xl.sheet_names}
    for cand in candidates:
        key = _norm(cand)
        if key in norm_map:
            return pd.read_excel(xl, norm_map[key])
    return None


def load_sources(xlsx_path=None, backlog_csv=None, avis_csv=None, complete_csv=None):
    """Charge les 3 DataFrames (backlog, avis, complete) depuis un xlsx OU 3 csv."""
    df_backlog = df_avis = df_complete = None
    if xlsx_path:
        xl = pd.ExcelFile(xlsx_path)
        df_backlog = _find_sheet(xl, "Backlog")
        df_avis = _find_sheet(xl, "Avis")
        df_complete = _find_sheet(xl, "Complété", "Complete", "Completé", "Termine", "Terminé")
    if backlog_csv is not None:
        df_backlog = pd.read_csv(backlog_csv)
    if avis_csv is not None:
        df_avis = pd.read_csv(avis_csv)
    if complete_csv is not None:
        df_complete = pd.read_csv(complete_csv)
    return df_backlog, df_avis, df_complete


def _extract_completed_games(df_complete):
    """Walks the 'Complété' sheet, handling year-separator rows and the
    summary columns (total hours, etc.) which are ignored."""
    if df_complete is None:
        return []
    games = []
    current_year = None
    jeu_col = next((c for c in df_complete.columns if _norm(c) == "jeu"), None)
    if jeu_col is None:
        return []
    # potential columns where a year marker might sit: either unnamed
    # columns from the original spreadsheet, or an explicit "Année" column if
    # re-importing a file already exported by the application itself.
    tail_cols = [c for c in df_complete.columns
                 if _norm(c).startswith("unnamed") or _norm(c) == "" or _norm(c) == "annee"]

    for _, row in df_complete.iterrows():
        title = row.get(jeu_col)
        title_empty = pd.isna(title) or str(title).strip() == ""

        # detect a year marker in one of the trailing columns
        year_found = None
        for tc in tail_cols:
            year_found = _as_year(row.get(tc))
            if year_found is not None:
                break

        if title_empty:
            if year_found:
                current_year = year_found
            continue  # separator / summary row, not a game

        def get(colname_variants):
            for c in df_complete.columns:
                if _norm(c) in colname_variants:
                    return row.get(c)
            return None

        # A finished game is assumed owned by default (you did play it):
        # if the sheet has no explicit value (the original spreadsheet has no
        # such column for finished games), default to "Yes". If an explicit
        # value exists (e.g. re-importing this app's own export, where the
        # user may have set "No"), it's respected.
        available_raw = _to_bool(get({"disponible"}))
        available = 1 if available_raw is None else available_raw

        games.append({
            "title": str(title).strip(),
            "dlc": _to_bool(get({"dlc"})) or 0,
            "abandoned": _to_bool(get({"abandonne", "abandonné", "abandon"})) or 0,
            "available": available,
            "hours_estimated": _to_float(get({"temps estime h", "temps estimes h", "temps estime en h"})),
            "hours_played": _to_float(get({"temps pris h"})),
            "duration_label": get({"duree"}),
            "month_finished": month_name_to_number(get({"mois fini"})),
            "year_finished": current_year,
            "worth_it": (str(get({"worth it"})).strip() if pd.notna(get({"worth it"})) else None),
            "rating": _to_float(get({"note", "note 10", "note 5"})),
        })
    return games


def _extract_backlog_games(df_backlog):
    if df_backlog is None:
        return []
    games = []
    jeu_col = next((c for c in df_backlog.columns if _norm(c) == "jeu"), None)
    if jeu_col is None:
        return []
    for _, row in df_backlog.iterrows():
        title = row.get(jeu_col)
        if pd.isna(title) or str(title).strip() == "":
            continue

        def get(colname_variants):
            for c in df_backlog.columns:
                if _norm(c) in colname_variants:
                    return row.get(c)
            return None

        games.append({
            "title": str(title).strip(),
            "hours_estimated": _to_float(get({"temps estimes en h", "temps estime en h", "temps estimes h"})),
            "available": _to_bool(get({"disponible"})),
            "status_label": get({"status"}),
        })
    return games


def _extract_reviews(df_avis):
    if df_avis is None:
        return []
    out = []
    jeu_col = next((c for c in df_avis.columns if _norm(c) in ("jeu", "jeux")), None)
    avis_col = next((c for c in df_avis.columns if _norm(c) in ("avis", "review")), None)
    if jeu_col is None:
        return []
    for _, row in df_avis.iterrows():
        title = row.get(jeu_col)
        if pd.isna(title) or str(title).strip() == "":
            continue
        review = row.get(avis_col) if avis_col else None
        out.append({
            "title": str(title).strip(),
            "review": (str(review).strip() if pd.notna(review) else None),
        })
    return out


def _richtext_to_markdown(value):
    """Convertit un texte enrichi openpyxl (gras/italique d'origine dans
    l'excel) en balises markdown, pour que la mise en forme d'origine ne soit
    pas perdue lors de l'import."""
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
    except ImportError:
        return value
    if not isinstance(value, CellRichText):
        return value
    parts = []
    for block in value:
        if isinstance(block, str):
            parts.append(block)
        elif isinstance(block, TextBlock):
            text = block.text
            font = block.font
            if font and getattr(font, "b", False) and getattr(font, "i", False):
                parts.append(f"***{text}***")
            elif font and getattr(font, "b", False):
                parts.append(f"**{text}**")
            elif font and getattr(font, "i", False):
                parts.append(f"*{text}*")
            else:
                parts.append(text)
    return "".join(parts)


def _extract_reviews_richtext(xlsx_path):
    """Re-reads the Avis sheet directly via openpyxl (rather than pandas) to
    recover the original bold/italic formatting. Returns None if it fails
    for any reason (falls back to the plain pandas extraction)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None
    try:
        wb = load_workbook(xlsx_path, rich_text=True, data_only=True)
    except Exception:
        return None

    ws = None
    for name in wb.sheetnames:
        if _norm(name) == "avis":
            ws = wb[name]
            break
    if ws is None:
        return None

    rows = list(ws.iter_rows())
    if not rows:
        return []
    header = [str(c.value).strip().lower() if c.value else "" for c in rows[0]]
    jeu_idx = next((i for i, h in enumerate(header) if h in ("jeu", "jeux")), None)
    avis_idx = next((i for i, h in enumerate(header) if h in ("avis", "review")), None)
    if jeu_idx is None:
        return None

    out = []
    for row in rows[1:]:
        title_cell = row[jeu_idx] if jeu_idx < len(row) else None
        title = title_cell.value if title_cell is not None else None
        if not title or not str(title).strip():
            continue
        review_val = row[avis_idx].value if avis_idx is not None and avis_idx < len(row) else None
        review_md = _richtext_to_markdown(review_val) if review_val else None
        out.append({"title": str(title).strip(), "review": review_md})
    return out


def extract_settings_from_xlsx(xlsx_path):
    """Reads the 'Settings' sheet (added to every export since Phase 1) and
    returns a dict of config values to merge back in, or {} if the sheet
    isn't present (e.g. importing an older export, or a hand-made file)."""
    try:
        xl = pd.ExcelFile(xlsx_path)
    except Exception:
        return {}
    df = _find_sheet(xl, "Settings", "Paramètres")
    if df is None or "Clé" not in df.columns or "Valeur" not in df.columns:
        return {}
    out = {}
    for _, row in df.iterrows():
        key = row.get("Clé")
        if pd.isna(key) or not str(key).strip():
            continue
        val = row.get("Valeur")
        out[str(key).strip()] = "" if pd.isna(val) else val
    return out


FUZZY_MIN_RATIO = 0.75    # below this, a candidate isn't worth suggesting at all
CONFIDENCE_MARGIN = 0.12  # min gap between the best and 2nd-best candidate to
                           # auto-suggest a single one; below that, titles are
                           # too close together (e.g. "(A)" vs "(B)" vs "(C)"
                           # variants of the same game) to guess safely —
                           # every plausible candidate is shown instead.
MAX_ALTERNATIVES = 5


def rank_review_candidates(review_title: str, completed_norm_to_id: dict):
    """Ranks every completed game by fuzzy similarity to `review_title` and
    classifies the result:
      - ("fuzzy", game_id, [])           one clearly-best candidate
      - ("ambiguous", game_id, [ids...]) several plausible candidates too
                                          close together to pick safely —
                                          game_id is the top-ranked one (used
                                          as the default selection), the rest
                                          are alternatives
      - (None, None, [])                 nothing above the similarity floor

    No substring shortcut: a plain substring hit (e.g. bare "Nier Automata"
    inside "Nier Automata ( Ending A )") used to override this ranking and
    could win over the far closer "Nier Automata (A)" simply for being
    shorter — that was the source of reviews landing on the wrong game
    variant. Ranking is by similarity alone now.
    """
    key = _norm(review_title)
    scored = []
    for norm_title, gid in completed_norm_to_id.items():
        if sequel_guard.sequel_conflict(key, norm_title):
            continue  # e.g. a "Dark Souls II" review must never suggest "Dark Souls III"
        ratio = difflib.SequenceMatcher(None, key, norm_title).ratio()
        if ratio >= FUZZY_MIN_RATIO:
            scored.append((gid, ratio))
    if not scored:
        return None, None, []

    scored.sort(key=lambda x: -x[1])
    best_id, best_ratio = scored[0]
    second_ratio = scored[1][1] if len(scored) > 1 else 0.0

    if best_ratio - second_ratio >= CONFIDENCE_MARGIN:
        return "fuzzy", best_id, []

    alt_ids = [gid for gid, _ in scored[1:MAX_ALTERNATIVES]]
    return "ambiguous", best_id, alt_ids


def import_all(xlsx_path=None, backlog_csv=None, avis_csv=None, complete_csv=None):
    """Imports the data into the database. Returns an import summary."""
    df_backlog, df_avis, df_complete = load_sources(xlsx_path, backlog_csv, avis_csv, complete_csv)

    completed = _extract_completed_games(df_complete)
    backlog = _extract_backlog_games(df_backlog)

    reviews = None
    if xlsx_path:
        reviews = _extract_reviews_richtext(xlsx_path)
    if reviews is None:
        reviews = _extract_reviews(df_avis)

    conn = get_conn()
    cur = conn.cursor()

    # index of completed games by normalized title, to match reviews
    completed_norm_to_id = {}

    for g in completed:
        cur.execute(
            """INSERT INTO games (title, status, dlc, abandoned, available, hours_estimated, hours_played,
               duration_label, month_finished, year_finished, worth_it, rating)
               VALUES (?, 'completed', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (g["title"], g["dlc"], g["abandoned"], g["available"], g["hours_estimated"], g["hours_played"],
             g["duration_label"], g["month_finished"], g["year_finished"], g["worth_it"], g["rating"]),
        )
        completed_norm_to_id[_norm(g["title"])] = cur.lastrowid

    for g in backlog:
        cur.execute(
            """INSERT INTO games (title, status, hours_estimated, available, notes)
               VALUES (?, 'backlog', ?, ?, ?)""",
            (g["title"], g["hours_estimated"], g["available"], g.get("status_label")),
        )

    matched, orphans, needs_confirmation = 0, 0, 0
    for r in reviews:
        key = _norm(r["title"])
        game_id = completed_norm_to_id.get(key)
        if game_id is not None:
            # exact normalized-title match: safe to apply automatically.
            cur.execute("UPDATE games SET review = ? WHERE id = ?", (r["review"], game_id))
            matched += 1
            continue

        # No exact match. Titles in 'Avis' don't always match 'Complété'
        # exactly (parentheses, wording like "Ending A" vs "(A)"...): rank
        # every completed game by fuzzy similarity and only *suggest* a
        # candidate — per the import-validation requirement, a suggestion is
        # never auto-applied. When several titles are nearly equally close
        # (e.g. "(A)"/"(B)"/"(C)" variants), all of them are surfaced as
        # alternatives instead of silently picking one (see
        # rank_review_candidates for why the old substring shortcut was
        # unsafe here).
        match_type, suggested_id, alt_ids = rank_review_candidates(r["title"], completed_norm_to_id)

        cur.execute(
            "INSERT INTO orphan_reviews (original_title, review, suggested_game_id, match_type, alternative_game_ids) "
            "VALUES (?, ?, ?, ?, ?)",
            (r["title"], r["review"], suggested_id, match_type, json.dumps(alt_ids) if alt_ids else None),
        )
        orphans += 1
        if suggested_id is not None:
            needs_confirmation += 1

    conn.commit()
    conn.close()

    return {
        "completed_imported": len(completed),
        "backlog_imported": len(backlog),
        "reviews_matched": matched,
        "reviews_orphan": orphans,
        "reviews_needs_confirmation": needs_confirmation,
    }
