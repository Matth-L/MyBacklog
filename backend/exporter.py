"""Exports the database data to CSV (3 files) or an Excel workbook (3+ sheets),
reusing the original 'My Backlog' file format, with formatting close to what
you'd do by hand (colored headers, year bands, bold/italic)."""
import io
import re
import zipfile
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from .db import get_conn, load_config, COVERS_DIR
from .dateutils import month_number_to_name

HEADER_FILL = PatternFill(start_color="FF7C5CFF", end_color="FF7C5CFF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)
YEAR_FILL = PatternFill(start_color="FF1F212B", end_color="FF1F212B", fill_type="solid")
YEAR_FONT = Font(color="FFC9FF5E", bold=True, size=12)


def _style_worksheet(ws, df):
    """Applies readable formatting: colored header, adjusted column widths,
    frozen pane, and distinct bands for year-separator rows."""
    if ws.max_row < 1 or not len(df.columns):
        return
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    for i, col in enumerate(df.columns, start=1):
        if len(df):
            content_len = max(len(str(v)) for v in df[col].astype(str).values)
        else:
            content_len = 0
        width = min(max(max(len(str(col)), content_len) + 2, 10), 50)
        ws.column_dimensions[get_column_letter(i)].width = width

    if "Jeu" in df.columns and "Année" in df.columns:
        jeu_idx = df.columns.get_loc("Jeu") + 1
        annee_idx = df.columns.get_loc("Année") + 1
        for row_i in range(2, ws.max_row + 1):
            jeu_val = ws.cell(row=row_i, column=jeu_idx).value
            annee_val = ws.cell(row=row_i, column=annee_idx).value
            if not jeu_val and annee_val:
                for cell in ws[row_i]:
                    cell.fill = YEAR_FILL
                    cell.font = YEAR_FONT


def _markdown_to_richtext(text):
    """Converts **bold** / *italic* markdown to Excel rich text, so the
    formatting stays close to the original once exported."""
    if not text or "*" not in text:
        return text
    tokens = re.split(r"(\*\*\*.+?\*\*\*|\*\*.+?\*\*|\*[^*]+?\*)", text)
    blocks = []
    for tok in tokens:
        if not tok:
            continue
        if tok.startswith("***") and tok.endswith("***") and len(tok) > 5:
            blocks.append(TextBlock(InlineFont(b=True, i=True), tok[3:-3]))
        elif tok.startswith("**") and tok.endswith("**") and len(tok) > 3:
            blocks.append(TextBlock(InlineFont(b=True), tok[2:-2]))
        elif tok.startswith("*") and tok.endswith("*") and len(tok) > 1:
            blocks.append(TextBlock(InlineFont(i=True), tok[1:-1]))
        else:
            blocks.append(tok)
    if len(blocks) == 1 and isinstance(blocks[0], str):
        return blocks[0]
    return CellRichText(*blocks)


def _apply_review_rich_text(ws, df):
    if "Avis" not in df.columns:
        return
    col_idx = df.columns.get_loc("Avis") + 1
    for row_i in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_i, column=col_idx)
        if isinstance(cell.value, str):
            rich = _markdown_to_richtext(cell.value)
            if isinstance(rich, CellRichText):
                cell.value = rich


def _fetch_games(status):
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM games WHERE status = ? ORDER BY year_finished IS NULL, year_finished, id",
        (status,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def build_info_df():
    """Small info sheet with the player's name, if set."""
    cfg = load_config()
    name = (cfg.get("player_name") or "").strip()
    rows = [{"Champ": "Exporté le", "Valeur": datetime.now().strftime("%Y-%m-%d %H:%M")}]
    if name:
        rows.insert(0, {"Champ": "Backlog de", "Valeur": name})
    return pd.DataFrame(rows)


# Config keys that make up "user settings" for export/import purposes. UI-only
# preferences (theme, language, dashboard layout, wallpaper...) live in the
# browser's localStorage and are intentionally not part of this cross-machine
# data bundle — they're device display preferences, not backlog data.
SETTINGS_EXPORT_KEYS = [
    "player_name",
    "rawg_api_key",
    "giantbomb_api_key",
    "internet_search_consent",
]


def build_settings_df():
    """User account/API settings, included in every export so a session
    (Excel data + settings) can be restored as a whole on another machine."""
    cfg = load_config()
    rows = [{"Clé": k, "Valeur": cfg.get(k, "")} for k in SETTINGS_EXPORT_KEYS]
    return pd.DataFrame(rows)


def settings_from_config() -> dict:
    """Settings subset as a plain dict, for the settings.json entry in csv
    export bundles."""
    cfg = load_config()
    return {k: cfg.get(k, "") for k in SETTINGS_EXPORT_KEYS}


# Excel/CSV Formula Injection: a game title or review imported from someone
# else's spreadsheet could be crafted to start with '=', '+', '-', or '@' —
# the characters every major spreadsheet app treats as "this cell is a
# formula". If that value is written back out unchanged (export, backup)
# and the resulting file is later opened in Excel, the "formula" can
# execute — this is the well-known OWASP-cataloged Excel/CSV Injection
# attack, and the exact reason a backlog imported from a friend's
# spreadsheet could otherwise smuggle a payload back out through this
# app's own exports. Prefixing a leading apostrophe is the standard
# mitigation: it's exactly what typing one manually does in Excel to force
# literal-text interpretation, and it round-trips harmlessly (openpyxl and
# pandas both display the text without the apostrophe).
_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _neutralize_formula_cell(value):
    if isinstance(value, str) and value.startswith(_FORMULA_TRIGGER_CHARS):
        return "'" + value
    return value


def _neutralize_formulas(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    return df.map(_neutralize_formula_cell)


def build_backlog_df():
    rows = _fetch_games("backlog")
    data = []
    for r in rows:
        data.append({
            "Jeu": r["title"],
            "Temps estimés (en h)": r["hours_estimated"],
            "Disponible ?": {1: "Yes", 0: "No"}.get(r["available"], None),
            "Status": r["notes"] or "",
        })
    if not data:
        return pd.DataFrame(columns=["Jeu", "Temps estimés (en h)", "Disponible ?", "Status"])
    return pd.DataFrame(data)


def build_avis_df():
    conn = get_conn()
    rows = conn.execute(
        "SELECT title, review FROM games WHERE status='completed' AND review IS NOT NULL AND review != ''"
    ).fetchall()
    orphans = conn.execute("SELECT original_title, review FROM orphan_reviews WHERE linked_game_id IS NULL").fetchall()
    conn.close()
    data = [{"Jeux": r["title"], "Avis": r["review"]} for r in rows]
    data += [{"Jeux": r["original_title"], "Avis": r["review"]} for r in orphans]
    if not data:
        return pd.DataFrame(columns=["Jeux", "Avis"])
    return pd.DataFrame(data)


def build_complete_df():
    rows = _fetch_games("completed")
    data = []
    current_year = None
    for r in rows:
        if r["year_finished"] != current_year:
            current_year = r["year_finished"]
            if current_year is not None:
                data.append({
                    "Jeu": None, "DLC ?": False, "Abandonné ?": False, "TEMPS ESTIME (h)": None,
                    "TEMPS PRIS (h)": None, "DUREE ?": None, "MOIS FINI": None, "WORTH IT?": None, "Note (/10)": None,
                    "Disponible ?": None, "Date exacte": None, "Année": current_year,
                })
        data.append({
            "Jeu": r["title"],
            "DLC ?": bool(r["dlc"]),
            "Abandonné ?": bool(r["abandoned"]),
            "TEMPS ESTIME (h)": r["hours_estimated"],
            "TEMPS PRIS (h)": r["hours_played"],
            "DUREE ?": r["duration_label"],
            "MOIS FINI": month_number_to_name(r["month_finished"]),
            "WORTH IT?": r["worth_it"],
            "Note (/10)": r["rating"],
            "Disponible ?": {1: "Yes", 0: "No"}.get(r["available"], None),
            "Date exacte": r["date_completed"] or "",
            "Année": None,
        })
    if not data:
        return pd.DataFrame(columns=["Jeu", "DLC ?", "Abandonné ?", "TEMPS ESTIME (h)", "TEMPS PRIS (h)",
                                      "DUREE ?", "MOIS FINI", "WORTH IT?", "Note (/10)",
                                      "Disponible ?", "Date exacte", "Année"])
    return pd.DataFrame(data)


def export_xlsx() -> bytes:
    buf = io.BytesIO()
    dfs = {
        "Info": build_info_df(),
        "Complété": build_complete_df(),
        "Avis": build_avis_df(),
        "Backlog": build_backlog_df(),
        "Settings": build_settings_df(),
    }
    dfs = {name: _neutralize_formulas(df) for name, df in dfs.items()}
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for name, df in dfs.items():
            df.to_excel(writer, sheet_name=name, index=False)
        for name, df in dfs.items():
            _style_worksheet(writer.sheets[name], df)
        _apply_review_rich_text(writer.sheets["Avis"], dfs["Avis"])
    buf.seek(0)
    return buf.read()


def export_csv_zip() -> bytes:
    import json
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("My_Backlog_-_Info.csv", _neutralize_formulas(build_info_df()).to_csv(index=False))
        zf.writestr("My_Backlog_-_Complete.csv", _neutralize_formulas(build_complete_df()).to_csv(index=False))
        zf.writestr("My_Backlog_-_Avis.csv", _neutralize_formulas(build_avis_df()).to_csv(index=False))
        zf.writestr("My_Backlog_-_Backlog.csv", _neutralize_formulas(build_backlog_df()).to_csv(index=False))
        zf.writestr("settings.json", json.dumps(settings_from_config(), ensure_ascii=False, indent=2))
    buf.seek(0)
    return buf.read()


def export_covers_zip() -> bytes:
    """Bundles every cover art file (jpg/png/webp/gif — anything actually
    on disk in the covers folder, not just what's currently referenced by a
    game) into a single downloadable zip, so the whole cover art collection
    can be backed up or moved independently of the game data itself."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(COVERS_DIR.iterdir()):
            if path.is_file():
                zf.write(path, arcname=path.name)
    buf.seek(0)
    return buf.read()
