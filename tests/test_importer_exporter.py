"""Tests unitaires pour l'import/export CSV & XLSX.

Objectif : garantir qu'aucune donnée n'est perdue lors d'un cycle
import -> export -> ré-import, y compris les cas particuliers du format
d'origine (lignes de séparation d'année, colonnes de résumé parasites,
avis qui ne correspondent à aucun jeu terminé).
"""
import io
import os
import sys
import json
import tempfile
from pathlib import Path

import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))


@pytest.fixture()
def temp_data_dir(monkeypatch, tmp_path):
    """Redirige la base de données vers un dossier temporaire pour chaque test."""
    monkeypatch.setenv("BACKLOG_DATA_DIR", str(tmp_path))
    # les modules backend lisent BACKLOG_DATA_DIR à l'import -> on doit les recharger
    for mod in list(sys.modules):
        if mod.startswith("backend"):
            del sys.modules[mod]
    from backend import db as db_module
    db_module.init_db()
    yield tmp_path


def _write_csv(tmp_path, name, df):
    path = tmp_path / name
    df.to_csv(path, index=False)
    return str(path)


def test_extract_backlog_games_basic(temp_data_dir):
    from backend.importer import _extract_backlog_games
    df = pd.DataFrame({
        "Jeu": ["Ace attorney", "Bioshock infinite", None],
        "Temps estimés (en h)": [None, 16, None],
        "Disponible ?": ["No", "Yes", None],
        "Status": [None, None, None],
    })
    games = _extract_backlog_games(df)
    assert len(games) == 2  # la ligne vide doit être ignorée
    assert games[0]["title"] == "Ace attorney"
    assert games[0]["available"] == 0
    assert games[1]["hours_estimated"] == 16.0
    assert games[1]["available"] == 1


def test_extract_completed_games_reads_note_out_of_10_with_half_points(temp_data_dir):
    """La colonne peut maintenant s'appeler 'Note (/10)' et contenir des
    valeurs décimales (ex: 7.5), suite au passage d'une échelle sur 5 à une
    échelle sur 10."""
    from backend.importer import _extract_completed_games
    df = pd.DataFrame({
        "Jeu": ["Parfait", "Moitie", "Nul"],
        "DLC ?": [False, False, False],
        "Note (/10)": [10.0, 7.5, 1.0],
    })
    games = _extract_completed_games(df)
    by_title = {g["title"]: g for g in games}
    assert by_title["Parfait"]["rating"] == 10.0
    assert by_title["Moitie"]["rating"] == 7.5
    assert by_title["Nul"]["rating"] == 1.0


def test_extract_completed_games_defaults_available_to_owned(temp_data_dir):
    """Un jeu fini est par défaut considéré comme possédé (le tableur
    d'origine n'a pas cette colonne pour les jeux finis) ; si une valeur
    explicite existe (ré-import d'un export de l'appli), elle est respectée."""
    from backend.importer import _extract_completed_games
    df = pd.DataFrame({
        "Jeu": ["Sans colonne Disponible", "Avec Non explicite", "Avec Oui explicite"],
        "DLC ?": [False, False, False],
        "NOTE": [5, 4, 3],
        "Disponible ?": [None, "No", "Yes"],
    })
    games = _extract_completed_games(df)
    by_title = {g["title"]: g for g in games}
    assert by_title["Sans colonne Disponible"]["available"] == 1
    assert by_title["Avec Non explicite"]["available"] == 0
    assert by_title["Avec Oui explicite"]["available"] == 1


def test_extract_completed_games_defaults_available_when_column_entirely_absent(temp_data_dir):
    """Idem, mais quand la colonne 'Disponible ?' n'existe même pas dans la
    feuille (cas du tableur d'origine avant l'ajout de cette fonctionnalité)."""
    from backend.importer import _extract_completed_games
    df = pd.DataFrame({"Jeu": ["Jeu ancien format"], "DLC ?": [False], "NOTE": [5]})
    games = _extract_completed_games(df)
    assert games[0]["available"] == 1


def test_extract_completed_games_handles_year_markers_and_summary_columns(temp_data_dir):
    """Reproduit la structure réelle : lignes vides marquant une année, et une
    colonne de fin de ligne qui contient parfois des libellés de résumé
    ('Nombre d'heure total :', etc.) qui ne doivent jamais être importés
    comme des jeux."""
    from backend.importer import _extract_completed_games
    df = pd.DataFrame({
        "Jeu": [None, "Persona 5 Royal", "Hollow Knight", None, "Elden Ring"],
        "DLC ?": [False, False, False, False, False],
        "TEMPS ESTIME (h)": [None, 123, 42, None, 100],
        "TEMPS PRIS (h)": [None, 100, 18, None, 83],
        "DUREE ?": [None, "1 an", "1m", None, "???"],
        "MOIS FINI": [None, "Janvier", "Mars", None, "Janvier"],
        "WORTH IT?": [None, "Yes", "Yes", None, "PEAK"],
        "NOTE": [0, 5, 3, 0, 5],
        "Unnamed: 8": [None] * 5,
        "Unnamed: 9": [2024, None, None, 2025, None],
    })
    games = _extract_completed_games(df)
    titles = [g["title"] for g in games]
    assert titles == ["Persona 5 Royal", "Hollow Knight", "Elden Ring"]
    assert games[0]["year_finished"] == 2024
    assert games[1]["year_finished"] == 2024
    assert games[2]["year_finished"] == 2025  # bascule bien après le 2e marqueur d'année


def test_extract_completed_games_recognizes_own_exported_annee_column(temp_data_dir):
    """Si on réimporte un fichier déjà exporté par l'application (colonne
    'Année' explicitement nommée, au lieu d'une colonne 'Unnamed' du tableur
    d'origine), les marqueurs d'année doivent toujours être reconnus."""
    from backend.importer import _extract_completed_games
    df = pd.DataFrame({
        "Jeu": [None, "Celeste", "Hades", None, "Outer Wilds"],
        "DLC ?": [False, False, False, False, False],
        "TEMPS ESTIME (h)": [None, 14, 30, None, 22],
        "TEMPS PRIS (h)": [None, 8, 25, None, 18],
        "DUREE ?": [None, "2d", "3w", None, "1w"],
        "MOIS FINI": [None, "Mars", "Juin", None, "Février"],
        "WORTH IT?": [None, "Yes", "Yes", None, "Yes"],
        "NOTE": [0, 5, 4, 0, 5],
        "Disponible ?": [None, "Yes", "Yes", None, "No"],
        "Année": [2024, None, None, 2025, None],
    })
    games = _extract_completed_games(df)
    titles = [g["title"] for g in games]
    assert titles == ["Celeste", "Hades", "Outer Wilds"]
    assert games[0]["year_finished"] == 2024
    assert games[1]["year_finished"] == 2024
    assert games[2]["year_finished"] == 2025


def test_extract_reviews_ignores_blank_rows(temp_data_dir):
    from backend.importer import _extract_reviews
    df = pd.DataFrame({"Jeux": ["Persona 5 royal", None], "Avis": ["Incroyable", None]})
    reviews = _extract_reviews(df)
    assert len(reviews) == 1
    assert reviews[0]["title"] == "Persona 5 royal"


def test_import_all_matches_reviews_with_slightly_different_titles(temp_data_dir):
    """Les titres dans 'Avis' ne correspondent pas toujours exactement à ceux de
    'Complété' (parenthèses, espaces...). Le matching approché doit s'en sortir,
    et ce qui ne matche vraiment pas doit finir en avis orphelin (pas perdu)."""
    from backend.importer import import_all
    from backend.db import get_conn

    backlog_df = pd.DataFrame({"Jeu": ["Outer wilds"], "Temps estimés (en h)": [22], "Disponible ?": ["No"]})
    complete_df = pd.DataFrame({
        "Jeu": ["Nier Automata ( Ending A )"],
        "DLC ?": [True], "TEMPS ESTIME (h)": [None], "TEMPS PRIS (h)": [10],
        "DUREE ?": ["2d"], "MOIS FINI": ["Janvier"], "WORTH IT?": ["Yes"], "NOTE": [5],
    })
    avis_df = pd.DataFrame({
        "Jeux": ["Nier Automata (A)", "Un jeu qui n'existe pas dans Complété"],
        "Avis": ["Super jeu", "Avis orphelin"],
    })

    with tempfile.TemporaryDirectory() as tmp:
        bpath = _write_csv(Path(tmp), "b.csv", backlog_df)
        cpath = _write_csv(Path(tmp), "c.csv", complete_df)
        apath = _write_csv(Path(tmp), "a.csv", avis_df)
        summary = import_all(backlog_csv=bpath, avis_csv=apath, complete_csv=cpath)

    assert summary["completed_imported"] == 1
    assert summary["backlog_imported"] == 1
    # Approximate matches (substring/fuzzy) are no longer auto-applied: they
    # become orphan reviews with a *suggested* game, awaiting confirmation.
    assert summary["reviews_matched"] == 0
    assert summary["reviews_orphan"] == 2
    assert summary["reviews_needs_confirmation"] == 1

    conn = get_conn()
    game = conn.execute("SELECT * FROM games WHERE status='completed'").fetchone()
    assert game["review"] is None  # not auto-modified

    orphans = {o["review"]: o for o in conn.execute("SELECT * FROM orphan_reviews").fetchall()}
    suggested = orphans["Super jeu"]
    assert suggested["suggested_game_id"] == game["id"]
    assert suggested["match_type"] == "fuzzy"

    unmatched = orphans["Avis orphelin"]
    assert unmatched["suggested_game_id"] is None
    assert unmatched["match_type"] is None
    conn.close()


def test_import_all_auto_applies_exact_title_matches(temp_data_dir):
    """Exact (normalized) title matches are safe and still applied
    automatically — only approximate matches require confirmation."""
    from backend.importer import import_all
    from backend.db import get_conn

    complete_df = pd.DataFrame({
        "Jeu": ["Hades"],
        "DLC ?": [False], "TEMPS ESTIME (h)": [None], "TEMPS PRIS (h)": [30],
        "DUREE ?": ["3d"], "MOIS FINI": ["Mars"], "WORTH IT?": ["Yes"], "NOTE": [9],
    })
    avis_df = pd.DataFrame({"Jeux": ["Hades"], "Avis": ["Excellent roguelike"]})

    with tempfile.TemporaryDirectory() as tmp:
        cpath = _write_csv(Path(tmp), "c.csv", complete_df)
        apath = _write_csv(Path(tmp), "a.csv", avis_df)
        summary = import_all(avis_csv=apath, complete_csv=cpath)

    assert summary["reviews_matched"] == 1
    assert summary["reviews_orphan"] == 0
    assert summary["reviews_needs_confirmation"] == 0

    conn = get_conn()
    game = conn.execute("SELECT * FROM games WHERE status='completed'").fetchone()
    assert game["review"] == "Excellent roguelike"
    conn.close()


def test_import_all_never_confuses_nier_automata_variants(temp_data_dir):
    """Regression test built from real reported data: five completed
    'Nier Automata' rows (bare title, plus (A)/(B)/(C)/(D&E) route variants)
    plus a 'Nier Replicant (A)'. Reviews with exact route wording ("(C)",
    "(D&E)") must exact-match their own variant. Reviews phrased differently
    ("Ending A", "Ending  B") must NOT silently land on the wrong variant —
    in particular, they must never fall through to the bare "Nier Automata"
    entry just because it happens to be a literal substring of the review
    title (the original bug: substring matching beat a far closer fuzzy
    match). Since the "(A)"/"(B)"/"(C)"/"(D&E)" candidates are all close in
    similarity to "Ending A"/"Ending B", these should come back ambiguous
    with alternatives listed, not a single silent guess."""
    from backend.importer import import_all
    from backend.db import get_conn

    complete_df = pd.DataFrame({
        "Jeu": ["Nier Automata (A)", "Nier Automata (B)", "Nier Automata (C)",
                "Nier Automata (D&E)", "Nier Automata", "Nier Replicant (A)"],
        "DLC ?": [False] * 6, "TEMPS ESTIME (h)": [None] * 6, "TEMPS PRIS (h)": [10] * 6,
        "DUREE ?": ["1d"] * 6, "MOIS FINI": ["Juin"] * 6, "WORTH IT?": ["Yes"] * 6, "NOTE": [9] * 6,
    })
    avis_df = pd.DataFrame({
        "Jeux": ["Nier Automata ( Ending A )", "Nier Automata ( Ending  B)",
                 "Nier Automata (C)", "Nier Automata (D&E)", "Nier Replicant (A)"],
        "Avis": ["Review A", "Review B", "Review C", "Review DE", "Review Replicant"],
    })

    with tempfile.TemporaryDirectory() as tmp:
        cpath = _write_csv(Path(tmp), "c.csv", complete_df)
        apath = _write_csv(Path(tmp), "a.csv", avis_df)
        summary = import_all(avis_csv=apath, complete_csv=cpath)

    conn = get_conn()
    games = {g["title"]: g for g in conn.execute("SELECT * FROM games").fetchall()}

    # Exact matches (C, D&E, Replicant A) apply immediately and correctly.
    assert summary["reviews_matched"] == 3
    assert games["Nier Automata (C)"]["review"] == "Review C"
    assert games["Nier Automata (D&E)"]["review"] == "Review DE"
    assert games["Nier Replicant (A)"]["review"] == "Review Replicant"

    # "Ending A" / "Ending B" are NOT auto-applied, and critically must not
    # have been silently assigned to the wrong game (bare "Nier Automata").
    assert games["Nier Automata"]["review"] is None
    assert games["Nier Automata (A)"]["review"] is None
    assert games["Nier Automata (B)"]["review"] is None

    orphans = {o["review"]: o for o in conn.execute("SELECT * FROM orphan_reviews").fetchall()}
    review_a = orphans["Review A"]
    review_b = orphans["Review B"]

    # The bare "Nier Automata" entry (a plain substring of both review
    # titles) must never be the chosen suggestion for either — that was
    # the exact bug being fixed.
    assert review_a["suggested_game_id"] != games["Nier Automata"]["id"]
    assert review_b["suggested_game_id"] != games["Nier Automata"]["id"]

    # With 4 very similar route-variant titles in play, this should be
    # flagged ambiguous rather than silently guessed, and the correct
    # variant must be among the choices offered.
    assert review_a["match_type"] == "ambiguous"
    alt_ids_a = json.loads(review_a["alternative_game_ids"])
    candidate_ids_a = [review_a["suggested_game_id"]] + alt_ids_a
    assert games["Nier Automata (A)"]["id"] in candidate_ids_a

    assert review_b["match_type"] == "ambiguous"
    alt_ids_b = json.loads(review_b["alternative_game_ids"])
    candidate_ids_b = [review_b["suggested_game_id"]] + alt_ids_b
    assert games["Nier Automata (B)"]["id"] in candidate_ids_b

    conn.close()


def test_import_all_flags_spec_example_as_needing_confirmation(temp_data_dir):
    """The exact scenario from the spec: 'Nier automata(A)' (completed) vs
    'Nier automata Ending A' (review) — plausibly the same game, but not an
    exact match, so it must be suggested for confirmation, not silently
    applied."""
    from backend.importer import import_all
    from backend.db import get_conn

    complete_df = pd.DataFrame({
        "Jeu": ["Nier automata(A)"],
        "DLC ?": [False], "TEMPS ESTIME (h)": [None], "TEMPS PRIS (h)": [40],
        "DUREE ?": ["4d"], "MOIS FINI": ["Juin"], "WORTH IT?": ["Yes"], "NOTE": [10],
    })
    avis_df = pd.DataFrame({"Jeux": ["Nier automata Ending A"], "Avis": ["Magnifique"]})

    with tempfile.TemporaryDirectory() as tmp:
        cpath = _write_csv(Path(tmp), "c.csv", complete_df)
        apath = _write_csv(Path(tmp), "a.csv", avis_df)
        summary = import_all(avis_csv=apath, complete_csv=cpath)

    assert summary["reviews_matched"] == 0
    assert summary["reviews_needs_confirmation"] == 1

    conn = get_conn()
    game = conn.execute("SELECT * FROM games WHERE status='completed'").fetchone()
    assert game["review"] is None
    orphan = conn.execute("SELECT * FROM orphan_reviews").fetchone()
    assert orphan["suggested_game_id"] == game["id"]
    conn.close()


def test_round_trip_no_data_loss(temp_data_dir):
    """Import -> export -> ré-import : le nombre de jeux et les valeurs clés
    doivent être identiques, rien ne doit disparaître silencieusement."""
    from backend.importer import import_all
    from backend import exporter
    from backend.db import get_conn

    backlog_df = pd.DataFrame({
        "Jeu": ["Black mesa", "Dishonored"],
        "Temps estimés (en h)": [18, 18],
        "Disponible ?": ["Yes", "Yes"],
    })
    complete_df = pd.DataFrame({
        "Jeu": [None, "Batman Arkham Asylum", "Celeste"],
        "DLC ?": [False, False, False],
        "TEMPS ESTIME (h)": [None, 16, 14],
        "TEMPS PRIS (h)": [None, 13, 6],
        "DUREE ?": [None, "1m", "2d"],
        "MOIS FINI": [None, "Janvier", "Aout"],
        "WORTH IT?": [None, "Yes", "Yes"],
        "NOTE": [0, 5, 4],
        "Unnamed: 9": [2024, None, None],
    })
    avis_df = pd.DataFrame({"Jeux": ["Batman Arkham Asylum"], "Avis": ["Excellent"]})

    with tempfile.TemporaryDirectory() as tmp:
        bpath = _write_csv(Path(tmp), "b.csv", backlog_df)
        cpath = _write_csv(Path(tmp), "c.csv", complete_df)
        apath = _write_csv(Path(tmp), "a.csv", avis_df)
        import_all(backlog_csv=bpath, avis_csv=apath, complete_csv=cpath)

    conn = get_conn()
    nb_before = conn.execute("SELECT COUNT(*) c FROM games").fetchone()["c"]
    conn.close()
    assert nb_before == 4  # 2 backlog + 2 completed

    xlsx_bytes = exporter.export_xlsx()
    assert len(xlsx_bytes) > 0

    # Ré-import à partir du fichier exporté, dans une base neuve
    xl = pd.ExcelFile(io.BytesIO(xlsx_bytes))
    assert set(xl.sheet_names) == {"Info", "Complété", "Avis", "Backlog", "Settings"}

    reimported_backlog = pd.read_excel(xl, "Backlog")
    reimported_complete = pd.read_excel(xl, "Complété")
    reimported_avis = pd.read_excel(xl, "Avis")

    assert len(reimported_backlog) == 2
    # 2 jeux + 1 ligne de séparation d'année = 3 lignes dans l'export Complété
    assert reimported_complete["Jeu"].notna().sum() == 2
    assert reimported_avis.iloc[0]["Avis"] == "Excellent"


def test_export_is_stable_with_empty_database(temp_data_dir):
    """Exporter une base vide ne doit jamais lever d'exception (cas du tout
    premier lancement sans import)."""
    from backend import exporter
    data = exporter.export_xlsx()
    assert len(data) > 0
    zdata = exporter.export_csv_zip()
    assert len(zdata) > 0


def test_info_sheet_includes_player_name_when_set(temp_data_dir):
    from backend import exporter
    from backend.db import save_config

    # sans nom renseigné : pas de ligne "Backlog de"
    df_empty = exporter.build_info_df()
    assert "Backlog de" not in df_empty["Champ"].values

    save_config({"configured": True, "player_name": "Alex"})
    df = exporter.build_info_df()
    row = df[df["Champ"] == "Backlog de"]
    assert len(row) == 1
    assert row.iloc[0]["Valeur"] == "Alex"


def test_export_applies_header_styling_and_year_band_highlight(temp_data_dir):
    """Vérifie que l'export n'est pas juste des données brutes : en-têtes
    colorés, largeurs de colonnes ajustées, et les lignes de séparation
    d'année mises en évidence (l'équivalent des 'bandes noires' de l'excel
    d'origine)."""
    from backend.db import get_conn
    from backend import exporter
    from openpyxl import load_workbook
    import io as io_module

    conn = get_conn()
    conn.execute(
        "INSERT INTO games (title, status, year_finished, rating) VALUES ('Jeu 2024', 'completed', 2024, 5)"
    )
    conn.commit()
    conn.close()

    xlsx_bytes = exporter.export_xlsx()
    wb = load_workbook(io_module.BytesIO(xlsx_bytes))
    ws = wb["Complété"]

    header_cell = ws.cell(row=1, column=1)
    assert header_cell.font.bold is True
    assert header_cell.fill.start_color.rgb == "FF7C5CFF"

    # trouve la ligne de séparation d'année (Jeu vide, Année renseignée) et
    # vérifie qu'elle est bien mise en évidence différemment des lignes de jeu
    year_col = [c.value for c in ws[1]].index("Année") + 1
    jeu_col = [c.value for c in ws[1]].index("Jeu") + 1
    found_band = False
    for row_i in range(2, ws.max_row + 1):
        if not ws.cell(row=row_i, column=jeu_col).value and ws.cell(row=row_i, column=year_col).value:
            found_band = True
            assert ws.cell(row=row_i, column=1).fill.start_color.rgb == "FF1F212B"
    assert found_band is True


def test_markdown_bold_italic_converted_to_excel_rich_text(temp_data_dir):
    from backend.db import get_conn
    from backend import exporter
    from openpyxl import load_workbook
    from openpyxl.cell.rich_text import CellRichText
    import io as io_module

    conn = get_conn()
    conn.execute(
        "INSERT INTO games (title, status, review) VALUES "
        "('Jeu markdown', 'completed', 'Un jeu **incroyable** et *original*.')"
    )
    conn.commit()
    conn.close()

    xlsx_bytes = exporter.export_xlsx()
    wb = load_workbook(io_module.BytesIO(xlsx_bytes), rich_text=True)
    ws = wb["Avis"]
    avis_col = [c.value for c in ws[1]].index("Avis") + 1
    cell_value = ws.cell(row=2, column=avis_col).value
    assert isinstance(cell_value, CellRichText)
    bold_blocks = [b for b in cell_value if hasattr(b, "font") and b.font and b.font.b]
    italic_blocks = [b for b in cell_value if hasattr(b, "font") and b.font and b.font.i and not b.font.b]
    assert any("incroyable" in b.text for b in bold_blocks)
    assert any("original" in b.text for b in italic_blocks)


def test_import_converts_excel_bold_italic_to_markdown(temp_data_dir, tmp_path):
    """Si l'excel d'origine avait du texte en gras/italique dans les avis,
    l'import doit le convertir en markdown plutôt que de le perdre."""
    from openpyxl import Workbook
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    from backend.importer import import_all
    from backend.db import get_conn

    wb = Workbook()
    ws_backlog = wb.active
    ws_backlog.title = "Backlog"
    ws_backlog.append(["Jeu", "Temps estimés (en h)", "Disponible ?"])

    ws_complete = wb.create_sheet("Complété")
    ws_complete.append(["Jeu", "DLC ?", "TEMPS ESTIME (h)", "TEMPS PRIS (h)", "DUREE ?", "MOIS FINI", "WORTH IT?", "NOTE"])
    ws_complete.append(["Jeu Riche", False, None, 10, "2d", "Mars", "Yes", 5])

    ws_avis = wb.create_sheet("Avis")
    ws_avis.append(["Jeux", "Avis"])
    rich = CellRichText(
        TextBlock(InlineFont(b=True), "Incroyable"), " et ",
        TextBlock(InlineFont(i=True), "original"), ".",
    )
    ws_avis.cell(row=2, column=1, value="Jeu Riche")
    ws_avis.cell(row=2, column=2, value=rich)

    xlsx_path = tmp_path / "test_richtext.xlsx"
    wb.save(xlsx_path)

    import_all(xlsx_path=str(xlsx_path))

    conn = get_conn()
    game = conn.execute("SELECT * FROM games WHERE title = 'Jeu Riche'").fetchone()
    conn.close()
    assert "**Incroyable**" in game["review"]
    assert "*original*" in game["review"]


def test_import_falls_back_to_plain_text_when_no_rich_formatting(temp_data_dir, tmp_path):
    """Un avis sans mise en forme particulière doit rester du texte simple,
    sans balises markdown parasites."""
    from openpyxl import Workbook
    from backend.importer import import_all
    from backend.db import get_conn

    wb = Workbook()
    ws_backlog = wb.active
    ws_backlog.title = "Backlog"
    ws_backlog.append(["Jeu", "Temps estimés (en h)", "Disponible ?"])

    ws_complete = wb.create_sheet("Complété")
    ws_complete.append(["Jeu", "DLC ?", "TEMPS ESTIME (h)", "TEMPS PRIS (h)", "DUREE ?", "MOIS FINI", "WORTH IT?", "NOTE"])
    ws_complete.append(["Jeu Simple", False, None, 5, "1d", "Avril", "Yes", 4])

    ws_avis = wb.create_sheet("Avis")
    ws_avis.append(["Jeux", "Avis"])
    ws_avis.append(["Jeu Simple", "Un avis tout à fait normal."])

    xlsx_path = tmp_path / "test_plain.xlsx"
    wb.save(xlsx_path)

    import_all(xlsx_path=str(xlsx_path))

    conn = get_conn()
    game = conn.execute("SELECT * FROM games WHERE title = 'Jeu Simple'").fetchone()
    conn.close()
    assert game["review"] == "Un avis tout à fait normal."


def test_abandoned_field_round_trips_through_export_and_reimport(temp_data_dir):
    """Un jeu marqué comme abandonné doit rester marqué après un export puis
    un ré-import du fichier généré."""
    from backend.db import get_conn
    from backend import exporter
    from backend.importer import _extract_completed_games

    conn = get_conn()
    conn.execute(
        "INSERT INTO games (title, status, abandoned, rating) VALUES ('Jeu abandonné', 'completed', 1, 2)"
    )
    conn.execute(
        "INSERT INTO games (title, status, abandoned, rating) VALUES ('Jeu fini', 'completed', 0, 5)"
    )
    conn.commit()
    conn.close()

    df = exporter.build_complete_df()
    assert bool(df[df["Jeu"] == "Jeu abandonné"].iloc[0]["Abandonné ?"]) is True
    assert bool(df[df["Jeu"] == "Jeu fini"].iloc[0]["Abandonné ?"]) is False

    reimported = _extract_completed_games(df)
    by_title = {g["title"]: g for g in reimported}
    assert by_title["Jeu abandonné"]["abandoned"] == 1
    assert by_title["Jeu fini"]["abandoned"] == 0


# ------------------------------------------------------------- Security: Excel/CSV formula injection

def test_neutralize_formula_cell_prefixes_trigger_characters():
    from backend.exporter import _neutralize_formula_cell

    for trigger in ["=cmd|'/c calc'!A1", "+1+1", "-2+3", "@SUM(A1:A9)", "\ttabstart"]:
        result = _neutralize_formula_cell(trigger)
        assert result == "'" + trigger
        assert result[0] == "'"


def test_neutralize_formula_cell_leaves_normal_text_and_non_strings_untouched():
    from backend.exporter import _neutralize_formula_cell

    assert _neutralize_formula_cell("Hades") == "Hades"
    assert _neutralize_formula_cell("A Normal - Title") == "A Normal - Title"  # dash mid-string, not leading
    assert _neutralize_formula_cell(42) == 42
    assert _neutralize_formula_cell(3.5) == 3.5
    assert _neutralize_formula_cell(None) is None
    assert _neutralize_formula_cell(True) is True


def test_exported_xlsx_never_writes_a_live_formula_cell(temp_data_dir):
    """A title or review imported from someone else's spreadsheet could be
    crafted to look like a formula (Excel/CSV Injection). If that value
    were ever written back out as an actual formula cell, opening the
    exported file in Excel could execute it. Every cell in every exported
    sheet must come back as a plain string, never openpyxl's 'f' (formula)
    data type."""
    from backend.db import get_conn
    from backend import exporter
    import openpyxl
    import io as _io

    conn = get_conn()
    conn.execute(
        "INSERT INTO games (title, status, review, notes) VALUES (?, 'completed', ?, ?)",
        ('=cmd|"/c calc"!A1', '=HYPERLINK("http://evil.example/steal","click")', "+1+1"),
    )
    conn.commit()
    conn.close()

    xlsx_bytes = exporter.export_xlsx()
    wb = openpyxl.load_workbook(_io.BytesIO(xlsx_bytes))
    formula_cells = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    formula_cells.append((ws.title, cell.coordinate, cell.value))
    assert formula_cells == [], f"Live formula cells found in export: {formula_cells}"

    # The original text must still be fully recoverable (just prefixed),
    # not silently dropped or mangled.
    ws = wb["Complété"]
    assert ws["A2"].value == '\'=cmd|"/c calc"!A1'


def test_exported_csv_neutralizes_formula_injection(temp_data_dir):
    from backend.db import get_conn
    from backend import exporter
    import zipfile, io as _io

    conn = get_conn()
    conn.execute(
        "INSERT INTO games (title, status) VALUES (?, 'completed')",
        ('=cmd|"/c calc"!A1',),
    )
    conn.commit()
    conn.close()

    zip_bytes = exporter.export_csv_zip()
    zf = zipfile.ZipFile(_io.BytesIO(zip_bytes))
    content = zf.read("My_Backlog_-_Complete.csv").decode("utf-8")
    assert "'=cmd|" in content
    # The raw (unprefixed) formula string must never appear on its own.
    assert '"=cmd|' not in content
